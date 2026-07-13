#!/usr/bin/env python3
"""
Script chuyển đổi file PDF từ vựng tiếng Nhật sang Excel

── CHẾ ĐỘ 1 FILE ──────────────────────────────────────────────
  python pdf_to_vocab_xlsx.py <input.pdf> [output.xlsx]
  python pdf_to_vocab_xlsx.py <input.pdf> [output.xlsx] --course-id ID_socapn4 --course-name "Sơ cấp N4" --level N4

── CHẾ ĐỘ BATCH (quét cả thư mục) ─────────────────────────────
  python pdf_to_vocab_xlsx.py --batch <thư_mục_gốc> [--output-dir <thư_mục_xuất>]

  Tự động:
    - Quét ĐỆ QUY toàn bộ file PDF có tên chứa "Từ vựng" hoặc "Từ_vựng"
      trong mọi thư mục con (Chương 16, 17, 18...)
    - Xuất MỖI bài 1 file Excel riêng, đặt tên: "TuVung_<LessonTitle>.xlsx"
    - In LOG chi tiết theo từng file: số từ tìm được, cảnh báo nếu
      nghi ngờ thiếu sót (số từ quá ít, có entry rỗng, v.v.)
    - Cuối cùng in BÁO CÁO TỔNG HỢP: tổng số file, tổng số từ,
      danh sách file cần kiểm tra lại thủ công

Options (dùng chung cho cả 2 chế độ):
  --course-id     ID khoá học (default: ID_socapn4)
  --course-name   Tên khoá học (default: Sơ cấp N4)
  --level         Cấp độ JLPT (default: N4)

Options riêng cho chế độ batch:
  --output-dir    Thư mục lưu các file Excel xuất ra (default: cùng thư mục
                   với từng file PDF nguồn)
  --pattern       Từ khoá nhận diện tên file PDF từ vựng (default: "Từ vựng")
"""

import sys
import re
import argparse
import subprocess
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ===== ARGS =====
def parse_args():
    p = argparse.ArgumentParser(description="Convert Japanese vocab PDF → Excel")
    p.add_argument("input_pdf", nargs="?", help="File PDF đầu vào (chế độ 1 file)")
    p.add_argument("output_xlsx", nargs="?", help="File Excel đầu ra (chế độ 1 file)")
    p.add_argument("--course-id",    default="ID_socapn4")
    p.add_argument("--course-name",  default="Sơ cấp N4")
    p.add_argument("--lesson-id",    default="")
    p.add_argument("--lesson-title", default="")
    p.add_argument("--level",        default="N4")
    # Batch mode
    p.add_argument("--batch",        default=None, metavar="THU_MUC_GOC",
                    help="Bật chế độ batch — quét đệ quy toàn bộ PDF trong thư mục này")
    p.add_argument("--output-dir",   default=None,
                    help="Thư mục lưu Excel xuất ra (batch mode). Mặc định: cùng chỗ với PDF gốc")
    p.add_argument("--pattern",      default="Từ vựng",
                    help='Từ khoá nhận diện file PDF từ vựng trong tên file (default: "Từ vựng")')
    return p.parse_args()


# ===== EXTRACT TEXT =====
def extract_text(pdf_path: str) -> str:
    r = subprocess.run(["pdftotext", "-layout", pdf_path, "-"],
                       capture_output=True, text=True, encoding="utf-8")
    if r.returncode != 0:
        raise RuntimeError(f"pdftotext error: {r.stderr}")
    return r.stdout


# ===== DETECT LESSON =====
def detect_lesson(text: str, pdf_path: str):
    filename = Path(pdf_path).stem
    m = re.search(r'(\d{1,2}[AB]?)', filename)
    if m:
        raw = m.group(1)
        return f"ID_{raw.lower()}", raw
    return "ID_unknown", "Unknown"


# ===== PARSE VOCABULARY =====
def parse_vocab(text: str) -> list:
    """
    Mỗi entry có dạng cơ bản:
    <word>    （TYPE）    <reading>    <meaning>

    Một số entry bị PDF wrap nghĩa dài thành nhiều dòng, ví dụ:
        (dòng 1, chỉ có phần đầu meaning, KHÔNG có type)
    よろしい    （A い）    よろしい
        (dòng 3, phần meaning còn lại, KHÔNG có type)

    → Cần gộp dòng liền trước/sau vào entry chính nếu chúng
      không tự chứa type nhưng nằm sát cạnh entry.
    """
    # Chấp nhận cả ngoặc full-width （）và half-width ASCII ()
    # vì các file PDF khác nhau dùng không thống nhất
    TYPE_RE = re.compile(
        r'[（(](N/Adv\.|N/V\s*III|N/V\s*II|N/V\s*I|V\s*III|V\s*II|V\s*I|A\s*い|A\s*な|Adv\.|N)[）)]'
    )

    # Dòng có chữ Nhật/Việt thực sự (không phải dòng trống/toàn space)
    def is_meaningful(line: str) -> bool:
        return bool(line.strip())

    raw_lines = text.splitlines()
    results = []

    i = 0
    n = len(raw_lines)
    while i < n:
        line = raw_lines[i]
        stripped = line.strip()

        if not stripped:
            i += 1
            continue

        m = TYPE_RE.search(stripped)
        if not m:
            i += 1
            continue

        type_raw  = m.group(1).strip()
        type_pos  = m.start()
        after_pos = m.end()

        word = stripped[:type_pos].strip()
        rest = stripped[after_pos:].strip()

        if not word:
            i += 1
            continue

        # ---- Xử lý case đặc biệt Pattern B: dòng ĐỘC LẬP phía trên
        # (dạng "label{spaces}reading{spaces}meaning", có type riêng hoặc
        # không) mới là entry chính thật, còn dòng word+type hiện tại chỉ
        # là 1 dòng-con khác bị PDF chèn lồng vào giữa (ví dụ ～代:
        #   ～代                　　　〜だい　　tiền～      ← entry CHÍNH, không có type ở đây
        #   電気代    （N）       でんきだい   tiền điện    ← dòng đang xét, TƯỞNG là entry riêng
        #   水道代                すいどうだい  tiền nước   ← dòng-con khác
        # )
        INDEPENDENT_ENTRY_RE = re.compile(
            r'^([\u3040-\u30FF\u30A0-\u30FF\u4E00-\u9FFF～]+)'
            r'\s{2,}([\u3040-\u30FF\u30A0-\u30FFー〜]+)\s{2,}(.+)$'
        )

        pattern_b_triggered = False
        if i > 0:
            prev_raw = raw_lines[i - 1].strip()
            if prev_raw and not TYPE_RE.search(prev_raw):
                b_m = INDEPENDENT_ENTRY_RE.match(prev_raw)
                # Loại trừ dòng tiêu đề cột (言葉/アクセント/意味/表現...)
                is_header_row = any(
                    kw in prev_raw for kw in ['言葉', '表現', 'アクセント', '意味', '語彙']
                )
                if b_m and not is_header_row:
                    b_word    = b_m.group(1).strip()
                    b_reading = b_m.group(2).strip()
                    b_meaning = b_m.group(3).strip()
                    # Chỉ tin nếu word ở dòng trên KHÁC với word đang xét
                    # (tránh trùng lặp/nhận nhầm chính dòng hiện tại)
                    if b_word and b_word != word:
                        results.append({
                            "word":            b_word,
                            "reading":         b_reading,
                            "type":            normalize_type(type_raw),
                            "meaning":         b_meaning,
                            "example":         "",
                            "example_meaning": "",
                        })
                        pattern_b_triggered = True

        # ---- Xử lý case đặc biệt: entry có NHIỀU dòng con reading+meaning
        # liệt kê bao quanh (ví dụ ～名 có 4 dòng: ～めい/いちめい/にめい/さんめい)
        # Nhận diện: rest rỗng VÀ dòng trước có dạng "kana... spaces... text"
        # (một dòng con reading+meaning hoàn chỉnh, không phải mảnh rác)
        SUB_ENTRY_RE = re.compile(
            r'^([\u3040-\u30FF\u30A0-\u30FF\uFF00-\uFFEF～ー]+)\s{2,}(.+)$'
        )

        if not rest:
            # Tìm ngược lên các dòng liệt kê phía TRƯỚC dòng word+type
            first_sub_reading, first_sub_meaning = "", ""
            j = i - 1
            found_any = False
            while j >= 0:
                prev_line = raw_lines[j].strip()
                if not prev_line:
                    break
                if TYPE_RE.search(prev_line):
                    break
                sub_m = SUB_ENTRY_RE.match(prev_line)
                if sub_m:
                    # Ghi đè liên tục — dòng GẦN dòng type nhất (j lớn nhất
                    # trong nhóm) sẽ là dòng cuối cùng được set, nhưng ta
                    # cần dòng ĐẦU TIÊN của cả nhóm → duyệt tiếp lên trên
                    first_sub_reading = sub_m.group(1).strip()
                    first_sub_meaning = sub_m.group(2).strip()
                    found_any = True
                    j -= 1
                    continue
                break

            if found_any:
                reading = first_sub_reading
                full_meaning = first_sub_meaning

                results.append({
                    "word":            word,
                    "reading":         reading,
                    "type":            normalize_type(type_raw),
                    "meaning":         full_meaning,
                    "example":         "",
                    "example_meaning": "",
                })
                i += 1
                continue

        # ---- Xử lý wrap: kiểm tra dòng NGAY TRƯỚC đó ----
        # Nếu dòng trước không trống và không chứa type → đó là phần đầu
        # của meaning bị đẩy lên trước (do cột meaning quá dài)
        # BỎ QUA nếu Pattern B đã tiêu thụ dòng trước đó rồi (tránh lặp lại
        # nội dung của entry độc lập vào meaning của entry hiện tại)
        prefix_meaning = ""
        if i > 0 and not pattern_b_triggered:
            prev = raw_lines[i - 1].strip()
            if prev and not TYPE_RE.search(prev) and not is_word_like_line(prev):
                prefix_meaning = prev

        # ---- Xử lý wrap: kiểm tra dòng NGAY SAU đó ----
        suffix_meaning = ""
        trailing_independent_entry = None
        if i + 1 < n:
            nxt = raw_lines[i + 1].strip()
            if pattern_b_triggered:
                # Nếu Pattern B đã trigger (dòng hiện tại nằm giữa 1 nhóm
                # ba entry độc lập), kiểm tra xem dòng SAU có phải 1 entry
                # độc lập khác không (ví dụ 水道代) — nếu đúng thì KHÔNG
                # ghép vào meaning, mà lưu riêng để thêm vào results sau
                nxt_m = INDEPENDENT_ENTRY_RE.match(nxt) if nxt else None
                if nxt_m and not TYPE_RE.search(nxt):
                    trailing_independent_entry = {
                        "word":    nxt_m.group(1).strip(),
                        "reading": nxt_m.group(2).strip(),
                        "meaning": nxt_m.group(3).strip(),
                    }
                    i += 1
            elif nxt and not TYPE_RE.search(nxt) and not is_word_like_line(nxt):
                suffix_meaning = nxt
                i += 1  # đã tiêu thụ dòng sau, nhảy qua

        if not rest and not prefix_meaning and not suffix_meaning:
            i += 1
            continue

        reading, meaning = split_reading_meaning(rest) if rest else ("", "")

        # Dọn rác: loại bỏ mảnh furigana kana đơn lẻ (1-2 ký tự) bị dính
        # do PDF canh cột ruby text lệch hàng (ví dụ: "よ" rác từ 良い）
        # Trường hợp 1: rác nằm cuối meaning có nội dung khác đứng trước
        meaning = re.sub(r'\s+[\u3040-\u30FF]{1,2}$', '', meaning).strip()
        # Trường hợp 2: meaning chỉ độc một mảnh kana rác, không có gì khác
        if re.fullmatch(r'[\u3040-\u30FF]{1,2}', meaning):
            meaning = ""

        # Ghép prefix/suffix vào meaning theo đúng thứ tự xuất hiện
        parts = [p for p in [prefix_meaning, meaning, suffix_meaning] if p]
        full_meaning = " ".join(parts).strip()
        full_meaning = re.sub(r'\s+', ' ', full_meaning)

        if not full_meaning:
            i += 1
            continue

        results.append({
            "word":            word,
            "reading":         reading,
            "type":            normalize_type(type_raw),
            "meaning":         full_meaning,
            "example":         "",
            "example_meaning": "",
        })

        # Thêm entry độc lập nằm ngay sau (nếu có) — ví dụ 水道代 đi kèm
        # cùng nhóm với ～代 / 電気代, dùng chung type với entry hiện tại
        if trailing_independent_entry:
            results.append({
                "word":            trailing_independent_entry["word"],
                "reading":         trailing_independent_entry["reading"],
                "type":            normalize_type(type_raw),
                "meaning":         trailing_independent_entry["meaning"],
                "example":         "",
                "example_meaning": "",
            })

        i += 1

    return results


def is_word_like_line(line: str) -> bool:
    """
    Kiểm tra xem dòng có 'giống' một dòng từ vựng mới không,
    hoặc là ruby/furigana rác (không phải phần meaning thật),
    để tránh bị nuốt nhầm vào meaning của entry liền kề.
    """
    line = line.strip()
    if not line:
        return False

    # Section headers / column headers — bỏ qua, không coi là meaning
    if any(kw in line for kw in ['言葉', '表現', 'アクセント', '意味', '語彙']):
        return True

    # Dòng ruby/furigana rác: chỉ gồm các mảnh kana ngắn (1-4 ký tự mỗi mảnh),
    # có thể tách nhau bởi khoảng trắng, không chứa ký tự Latin/Việt
    # (ví dụ: "わる", hoặc "ご    ご" — furigana của 語 bị lặp 2 vị trí)
    PURE_KANA_FRAGMENTS = re.compile(r'^[\u3040-\u30FF]{1,4}(\s+[\u3040-\u30FF]{1,4})*$')
    if PURE_KANA_FRAGMENTS.match(line) and len(line.replace(' ', '')) <= 8:
        return True

    return False


def split_reading_meaning(text: str):
    """
    Tách reading (JP) và meaning (Việt) từ phần sau type.

    Pattern reading có thể là:
      - "ほんもの"
      - "きっと"
      - "ちゅうし ( する )"   ← có ASCII () bọc する
      - "せいこう ( する )"
    """
    text = text.strip()

    # Pattern 1: reading có "( する )" hoặc "（する）"
    # JP_chars + optional whitespace + optional "( する )" + whitespace + meaning
    m = re.match(
        r'^([\u3040-\u30FF\u4E00-\u9FFF\uFF00-\uFFEF（）ー～\s]+'
        r'(?:[\(（]\s*する\s*[)）])?)'
        r'\s{2,}'   # ít nhất 2 space để phân cách với meaning
        r'(.+)$',
        text
    )
    if m:
        reading = re.sub(r'\s+', ' ', m.group(1)).strip()
        meaning = m.group(2).strip()
        return reading, meaning

    # Pattern 2: fallback — tìm ký tự Latin đầu tiên
    JP_SET = re.compile(r'[\u3040-\u30FF\u4E00-\u9FFF\uFF00-\uFFEF（）ー～\(\)\s]')
    split_idx = len(text)
    for i, ch in enumerate(text):
        if not JP_SET.match(ch):
            split_idx = i
            break

    reading = re.sub(r'\s+', ' ', text[:split_idx]).strip().rstrip('(').strip()
    meaning = text[split_idx:].strip()
    return reading, meaning


def normalize_type(raw: str) -> str:
    raw = re.sub(r'\s+', ' ', raw).strip()
    mapping = {
        'N/Adv.':  'N/Adv.',
        'N/V III': 'N/V III',
        'N/V II':  'N/V II',
        'N/V I':   'N/V I',
        'V III':   'V III',
        'V II':    'V II',
        'V I':     'V I',
        'A い':    'A い',
        'A な':    'A な',
        'Adv.':    'Adv.',
        'N':       'N',
    }
    for key, val in mapping.items():
        if key == raw:
            return val
    return raw


# ===== XỬ LÝ 1 FILE (dùng chung cho single-mode và batch-mode) =====
def process_one_pdf(pdf_path: Path, output_path: Path,
                     course_id: str, course_name: str, level: str,
                     lesson_id: str = "", lesson_title: str = "",
                     verbose: bool = True) -> dict:
    """
    Xử lý 1 file PDF → xuất Excel.
    Trả về dict kết quả để batch mode tổng hợp log/báo cáo, gồm:
      status: "ok" | "warning" | "error"
      word_count, warnings (list các cảnh báo cụ thể), error_message
    """
    result = {
        "pdf_path":     str(pdf_path),
        "output_path":  str(output_path),
        "lesson_title": lesson_title,
        "status":       "ok",
        "word_count":   0,
        "warnings":     [],
        "error_message": "",
    }

    try:
        text = extract_text(str(pdf_path))
    except Exception as e:
        result["status"] = "error"
        result["error_message"] = f"Không đọc được PDF: {e}"
        return result

    # Detect lesson nếu chưa truyền sẵn
    if not lesson_id or not lesson_title:
        det_id, det_title = detect_lesson(text, str(pdf_path))
        lesson_id    = lesson_id    or det_id
        lesson_title = lesson_title or det_title
        result["lesson_title"] = lesson_title

    try:
        vocab = parse_vocab(text)
    except Exception as e:
        result["status"] = "error"
        result["error_message"] = f"Lỗi khi parse: {e}"
        return result

    result["word_count"] = len(vocab)

    if not vocab:
        result["status"] = "error"
        result["error_message"] = "Không parse được từ nào (0 từ)"
        return result

    # ---- Đánh giá "confidence" — heuristic phát hiện file nghi ngờ lỗi ----
    # (KHÔNG sửa dữ liệu, chỉ CẢNH BÁO để người dùng tự kiểm tra lại)
    warnings = check_vocab_quality(vocab, text)
    result["warnings"] = warnings
    if warnings:
        result["status"] = "warning"

    # Xuất Excel
    try:
        create_excel(vocab, str(output_path),
                     course_id=course_id, course_name=course_name,
                     lesson_id=lesson_id, lesson_title=lesson_title, level=level)
    except Exception as e:
        result["status"] = "error"
        result["error_message"] = f"Lỗi khi ghi Excel: {e}"
        return result

    if verbose:
        print(f"📖 {pdf_path.name}")
        print(f"   📚 Bài: {lesson_title}  (ID: {lesson_id})")
        print(f"   📝 {len(vocab)} từ vựng")
        for w in warnings:
            print(f"   ⚠️  {w}")
        print(f"   ✅ Đã lưu: {output_path}\n")

    return result


def check_vocab_quality(vocab: list, raw_text: str) -> list:
    """
    Heuristic KIỂM TRA (không sửa dữ liệu) để phát hiện dấu hiệu file
    có thể bị parse thiếu/sai, dựa trên kinh nghiệm từ các file đã test
    thủ công (15A, 17A, 23B, 25B, 29A, 33B):

      1. Số từ tìm được quá ít so với số dòng có type-pattern trong PDF
         → có thể còn sót entry dạng phức tạp (như 翻訳（する）ở 23B,
           hoặc layout "khiêm nhường ngữ" nhiều tầng như 33B)
      2. Entry có meaning/reading rỗng hoặc quá ngắn bất thường (1-2 ký tự)
         → dấu hiệu bị dính rác hoặc cắt cụt (như trường hợp よろしい
           trước khi được fix)
      3. Entry có ký tự rác điển hình còn sót (chuỗi kana lặp lại y hệt
         2 lần trong 1 meaning — dấu hiệu chưa xử lý được reading dạng
         "nằm 2 dòng bao quanh" như アイデア/チャレンジ ở 25B)
    """
    warnings = []

    # (1) So sánh số entry parse được với số dòng chứa type-pattern thô
    TYPE_RE_CHECK = re.compile(
        r'[（(](N/Adv\.|N/V\s*III|N/V\s*II|N/V\s*I|V\s*III|V\s*II|V\s*I|A\s*い|A\s*な|Adv\.|N)[）)]'
    )
    raw_type_lines = sum(1 for l in raw_text.splitlines() if TYPE_RE_CHECK.search(l))
    if raw_type_lines > 0 and len(vocab) < raw_type_lines * 0.7:
        warnings.append(
            f"Số từ parse được ({len(vocab)}) thấp hơn nhiều so với số dòng "
            f"có type-pattern trong PDF ({raw_type_lines}) — có thể còn sót entry phức tạp"
        )

    # (2) Entry có meaning/reading bất thường
    short_meaning_count = 0
    empty_reading_count = 0
    for v in vocab:
        if len(v["meaning"].strip()) <= 1:
            short_meaning_count += 1
        if not v["reading"].strip():
            empty_reading_count += 1

    if short_meaning_count > 0:
        warnings.append(
            f"{short_meaning_count} từ có nghĩa quá ngắn (≤1 ký tự) — nghi ngờ bị cắt cụt"
        )
    if empty_reading_count > 0:
        warnings.append(
            f"{empty_reading_count} từ THIẾU reading — nghi ngờ reading nằm ở dòng "
            f"tách biệt bao quanh mà script chưa gộp được"
        )

    # (3) Meaning có dấu hiệu lặp rác (ví dụ "チャレンジ（する）thử sức チャレンジ ( する )")
    for v in vocab:
        words_in_meaning = re.findall(r'[\u3040-\u30FF\u30A0-\u30FF\u4E00-\u9FFF]{3,}', v["meaning"])
        if len(words_in_meaning) != len(set(words_in_meaning)):
            warnings.append(
                f"Từ '{v['word']}' có nghĩa nghi ngờ dính rác lặp lại: \"{v['meaning'][:60]}...\""
            )

    return warnings


# ===== CREATE EXCEL =====
def create_excel(vocab: list, out: str, course_id: str, course_name: str,
                 lesson_id: str, lesson_title: str, level: str):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Từ vựng {lesson_title}"[:31]  # Excel title max 31 chars

    # Styles
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", start_color="1565C0")
    dat_font  = Font(name="Arial", size=11)
    even_fill = PatternFill("solid", start_color="E3F2FD")
    odd_fill  = PatternFill("solid", start_color="FFFFFF")
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="BDBDBD")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Course ID", "Course Name", "Lesson ID", "Lesson Title",
        "Word", "Reading", "Type", "Meaning",
        "Example", "Example Meaning", "Level"
    ]

    # Header row
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = bdr
    ws.row_dimensions[1].height = 28

    # Data rows
    CENTER_COLS = {1, 3, 4, 7, 11}
    for ri, v in enumerate(vocab, 2):
        fill = even_fill if ri % 2 == 0 else odd_fill
        row_vals = [
            course_id, course_name, lesson_id, lesson_title,
            v["word"], v["reading"], v["type"], v["meaning"],
            v["example"], v["example_meaning"], level
        ]
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = dat_font
            cell.fill      = fill
            cell.border    = bdr
            cell.alignment = center if ci in CENTER_COLS else left
        ws.row_dimensions[ri].height = 22

    # Column widths
    widths = [15, 14, 10, 13, 22, 22, 10, 35, 35, 35, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{len(vocab)+1}"

    wb.save(out)


# ===== BATCH MODE =====
def find_vocab_pdfs(root_dir: Path, pattern: str) -> list:
    """
    Quét ĐỆ QUY toàn bộ thư mục con, tìm file .pdf có tên chứa `pattern`
    (mặc định "Từ vựng"). Trả về danh sách Path đã sắp xếp theo tên
    để log dễ theo dõi.
    """
    matches = []
    for pdf_file in root_dir.rglob("*.pdf"):
        if pattern.lower().replace(" ", "") in pdf_file.stem.lower().replace(" ", "").replace("_", ""):
            matches.append(pdf_file)
    return sorted(matches, key=lambda p: str(p))


def run_batch(root_dir: str, output_dir: str, course_id: str,
              course_name: str, level: str, pattern: str):
    root = Path(root_dir)
    if not root.exists():
        print(f"❌ Thư mục không tồn tại: {root_dir}")
        sys.exit(1)

    out_dir = Path(output_dir) if output_dir else None
    if out_dir:
        out_dir.mkdir(parents=True, exist_ok=True)

    pdf_files = find_vocab_pdfs(root, pattern)

    print("=" * 70)
    print(f"🔍 QUÉT THƯ MỤC: {root_dir}")
    print(f"   Từ khoá nhận diện: \"{pattern}\"")
    print(f"   Tìm thấy: {len(pdf_files)} file PDF từ vựng")
    print("=" * 70)
    print()

    if not pdf_files:
        print("⚠️  Không tìm thấy file PDF nào khớp từ khoá. Kiểm tra lại --pattern.")
        sys.exit(1)

    all_results = []
    for pdf_path in pdf_files:
        if out_dir:
            output_path = out_dir / f"TuVung_{pdf_path.stem.replace('Từ_vựng_', '').replace('Từ vựng ', '').strip()}.xlsx"
        else:
            output_path = pdf_path.with_name(
                f"TuVung_{pdf_path.stem.replace('Từ_vựng_', '').replace('Từ vựng ', '').strip()}.xlsx"
            )

        result = process_one_pdf(
            pdf_path, output_path,
            course_id=course_id, course_name=course_name, level=level,
            verbose=True,
        )
        all_results.append(result)

    print_batch_summary(all_results)
    write_batch_log_file(all_results, root)


def print_batch_summary(results: list):
    total_files = len(results)
    ok_files       = [r for r in results if r["status"] == "ok"]
    warning_files  = [r for r in results if r["status"] == "warning"]
    error_files    = [r for r in results if r["status"] == "error"]
    total_words    = sum(r["word_count"] for r in results)

    print("=" * 70)
    print("📊 BÁO CÁO TỔNG HỢP")
    print("=" * 70)
    print(f"Tổng số file PDF     : {total_files}")
    print(f"  ✅ Chạy tốt (không cảnh báo) : {len(ok_files)}")
    print(f"  ⚠️  Có cảnh báo (nên kiểm tra) : {len(warning_files)}")
    print(f"  ❌ Lỗi (không xuất được)      : {len(error_files)}")
    print(f"Tổng số từ vựng đã trích xuất : {total_words}")
    print()

    if warning_files:
        print("── DANH SÁCH FILE CÓ CẢNH BÁO (nên kiểm tra lại thủ công) ──────")
        for r in warning_files:
            print(f"  📄 {Path(r['pdf_path']).name}  ({r['word_count']} từ)")
            for w in r["warnings"]:
                print(f"      ⚠️  {w}")
        print()

    if error_files:
        print("── DANH SÁCH FILE LỖI (KHÔNG xuất được Excel) ──────────────────")
        for r in error_files:
            print(f"  📄 {Path(r['pdf_path']).name}")
            print(f"      ❌ {r['error_message']}")
        print()

    print("=" * 70)


def write_batch_log_file(results: list, root_dir: Path):
    """Ghi log chi tiết ra file .txt để xem lại sau, kèm timestamp."""
    log_path = root_dir / f"batch_convert_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

    total_files = len(results)
    ok_files      = [r for r in results if r["status"] == "ok"]
    warning_files = [r for r in results if r["status"] == "warning"]
    error_files   = [r for r in results if r["status"] == "error"]
    total_words   = sum(r["word_count"] for r in results)

    with open(log_path, "w", encoding="utf-8") as f:
        f.write(f"BATCH CONVERT LOG — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 70 + "\n")
        f.write(f"Tổng số file PDF: {total_files}\n")
        f.write(f"  OK (không cảnh báo)  : {len(ok_files)}\n")
        f.write(f"  WARNING (nên kiểm tra): {len(warning_files)}\n")
        f.write(f"  ERROR (không xuất được): {len(error_files)}\n")
        f.write(f"Tổng số từ vựng: {total_words}\n")
        f.write("=" * 70 + "\n\n")

        for r in results:
            f.write(f"[{r['status'].upper()}] {r['pdf_path']}\n")
            f.write(f"    → Output: {r['output_path']}\n")
            f.write(f"    → Bài: {r['lesson_title']}  |  Số từ: {r['word_count']}\n")
            if r["warnings"]:
                for w in r["warnings"]:
                    f.write(f"    ⚠️  {w}\n")
            if r["error_message"]:
                f.write(f"    ❌ {r['error_message']}\n")
            f.write("\n")

    print(f"📄 Log chi tiết đã lưu tại: {log_path}")


# ===== MAIN =====
def main():
    args = parse_args()

    # ---- CHẾ ĐỘ BATCH ----
    if args.batch:
        run_batch(
            root_dir=args.batch,
            output_dir=args.output_dir,
            course_id=args.course_id,
            course_name=args.course_name,
            level=args.level,
            pattern=args.pattern,
        )
        return

    # ---- CHẾ ĐỘ 1 FILE (giữ nguyên hành vi cũ) ----
    if not args.input_pdf:
        print("❌ Thiếu tham số. Dùng --batch <thư_mục> để chạy hàng loạt,")
        print("   hoặc truyền input_pdf để chạy 1 file.")
        sys.exit(1)

    if not Path(args.input_pdf).exists():
        print(f"❌ File không tồn tại: {args.input_pdf}")
        sys.exit(1)

    out = args.output_xlsx or str(Path(args.input_pdf).with_suffix('.xlsx'))

    print(f"📖 Đọc PDF: {args.input_pdf}")
    text = extract_text(args.input_pdf)

    # Detect lesson
    lesson_id    = args.lesson_id
    lesson_title = args.lesson_title
    if not lesson_id or not lesson_title:
        det_id, det_title = detect_lesson(text, args.input_pdf)
        lesson_id    = lesson_id    or det_id
        lesson_title = lesson_title or det_title

    print(f"📚 Bài: {lesson_title}  (ID: {lesson_id})")

    vocab = parse_vocab(text)
    print(f"📝 Tìm thấy: {len(vocab)} từ vựng")

    if not vocab:
        print("⚠️  Không parse được từ nào. Kiểm tra format PDF.")
        sys.exit(1)

    # Cảnh báo chất lượng (không sửa dữ liệu, chỉ hiển thị)
    warnings = check_vocab_quality(vocab, text)
    if warnings:
        print("\n⚠️  CẢNH BÁO — nên kiểm tra lại các từ sau:")
        for w in warnings:
            print(f"   • {w}")

    # Preview
    print("\n── Preview ──────────────────────────────────────────────────")
    for v in vocab:
        print(f"  {v['word']:<22} {v['reading']:<22} {v['type']:<10} {v['meaning']}")
    print("─────────────────────────────────────────────────────────────\n")

    create_excel(vocab, out,
                 course_id=args.course_id, course_name=args.course_name,
                 lesson_id=lesson_id, lesson_title=lesson_title, level=args.level)

    print(f"✅ Đã lưu: {out}")


if __name__ == "__main__":
    main()